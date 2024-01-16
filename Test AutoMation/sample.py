from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

def get_longest_shortest_options(search_results):
    if not search_results:
        return None, None

    longest_option = max(search_results, key=lambda x: len(x.text))
    shortest_option = min(search_results, key=lambda x: len(x.text))

    return longest_option.text, shortest_option.text


excel_file_path = "C:/Users/rahim/OneDrive/Desktop/Test AutoMation/Excel.xlsx"
chrome_driver_path = "c:/Program Files (x86)/chromedriver-win64/chromedriver.exe"

workbook = openpyxl.load_workbook(excel_file_path)

options = webdriver.ChromeOptions()
options.add_argument(chrome_driver_path)
driver = webdriver.Chrome(options=options)

for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    keyboard_names = [sheet.cell(row=i, column=3).value for i in range(3, 13)]

    for keyboard_name in keyboard_names:
        try:
            driver.get("https://www.google.com/")

            search_box = driver.find_element("name", "q")
            search_box.send_keys(keyboard_name)

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//ul[@role='listbox']/li"))
            )

            search_results = driver.find_elements(By.XPATH, "//ul[@role='listbox']/li")
            longest_option, shortest_option = get_longest_shortest_options(search_results)
            row_index = keyboard_names.index(keyboard_name) + 3
            sheet.cell(row=row_index, column=4).value = longest_option
            sheet.cell(row=row_index, column=5).value = shortest_option

        except Exception as e:
            print(f"Error processing {keyboard_name} in sheet {sheet_name}: {str(e)}")

workbook.save(excel_file_path)

driver.quit()
